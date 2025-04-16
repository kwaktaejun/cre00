
from flask import Flask, render_template, request, send_file
from openpyxl import Workbook
from openpyxl.styles import Alignment, PatternFill, Border, Side
from io import BytesIO
from datetime import datetime

app = Flask(__name__)

@app.route('/', methods=['GET', 'POST'])
def index():
    if request.method == 'POST':
        table_data = request.form.get('tableData')
        memo = request.form.get('memo')
        rows = [row.split("\t") for row in table_data.strip().split("\n") if row.strip()]

        wb = Workbook()
        ws = wb.active
        ws.title = "업로드 데이터"

        headers = ["구분", "계약여부", "식별번호", "계약금액", "제품모델명", "품명", "모델명", "규격", "수량", "원산지", "구성종류", "제품원가", "원천제조사", "수익률", "비고", "메모"]
        ws.append(headers)

        # 스타일 요소
        thin_border = Border(left=Side(style='thin', color='FFFFFF'), right=Side(style='thin', color='FFFFFF'),
                             top=Side(style='thin', color='FFFFFF'), bottom=Side(style='thin', color='FFFFFF'))
        align_center = Alignment(horizontal='center', vertical='center')
        fill_blue = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")
        fill_white = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

        color_map = {}
        current_fill = None
        model_index = headers.index("제품모델명")

        for row_idx, row in enumerate(rows, start=2):
            full_row = row + [""] * (16 - len(row))
            ws.append(full_row)

            model_name = full_row[model_index]
            if model_name in color_map:
                current_fill = color_map[model_name]
            else:
                current_fill = fill_blue if (len(color_map) % 2 == 0) else fill_white
                color_map[model_name] = current_fill

            for col_idx in range(1, 17):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.border = thin_border
                cell.alignment = align_center
                cell.fill = current_fill

            # 수익률 수식 입력 (14번째 인덱스: 수익률)
            try:
                contract_price = float(str(full_row[3]).replace(",", ""))
                cost_price = float(str(full_row[11]).replace(",", ""))
                formula = f'=IF(D{row_idx}="", "", ROUND((D{row_idx}-L{row_idx}-(D{row_idx}*0.45))/D{row_idx}*100, 2))'
                ws.cell(row=row_idx, column=14).value = formula
            except:
                pass

        # 열 너비 설정
        col_widths = [10, 10, 15, 15, 20, 15, 20, 25, 8, 10, 12, 15, 15, 10, 12, 20]
        for i, width in enumerate(col_widths, start=1):
            ws.column_dimensions[chr(64 + i if i <= 26 else 64 + i // 26) + chr(64 + i % 26)].width = width

        # 메모 시트
        memo_ws = wb.create_sheet("업로드 메모")
        memo_ws.append(["업로드 메모"])
        memo_ws.append([memo])

        file_stream = BytesIO()
        wb.save(file_stream)
        file_stream.seek(0)

        filename = f"마스터시트_업로드_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        return send_file(file_stream, as_attachment=True, download_name=filename, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

    return render_template('index.html')
